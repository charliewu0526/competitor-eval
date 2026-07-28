"""T2 pilot task: sum an Excel column range into one cell, save.

Domain #1 (closed-source desktop app), office-suite. Unlike T1 (WeChat, whose
end-state a script can't read -> human-verified flags), T2 produces a FILE
artifact. So the primary end-state IS machine-checkable: open the saved workbook
and verify C14 == sum(C2:C13). No intern self-report on the core goal.

The machine assertion reads ctx['artifact_path'] — the server-side saved
artifact (authoritative, populated by intake._build_ctx). If the file is missing
/ unreadable / the cell is wrong, it fails (未验证 != 通过, no fabricated pass).
"""
import pathlib
from pipeline import objective as O
from pipeline import taskbank as TB

TASK_DIR = pathlib.Path(__file__).resolve().parent / "T2-excel-sum-001"
TASK = TB.assert_valid(TASK_DIR)

# The range the task specifies: sum of C2:C13 lands in C14, sheet 'Q1'.
_SHEET = "Q1"
_SUM_ROWS = range(2, 14)   # C2..C13 inclusive
_TARGET_CELL = "C14"


def _read_c14_and_sum(artifact_path):
    """Open the saved workbook -> (c14_value, computed_sum). Raise on any problem
    so the assertion treats it as a fail (can't verify -> not passed)."""
    import openpyxl
    wb = openpyxl.load_workbook(artifact_path, data_only=True)
    ws = wb[_SHEET] if _SHEET in wb.sheetnames else wb.active
    total = 0
    for r in _SUM_ROWS:
        v = ws[f"C{r}"].value
        total += v if isinstance(v, (int, float)) else 0
    c14 = ws[_TARGET_CELL].value
    return c14, total


def _c14_equals_sum(ctx: dict) -> bool:
    p = ctx.get("artifact_path")
    if not p or not pathlib.Path(p).expanduser().is_file():
        return False
    try:
        c14, total = _read_c14_and_sum(pathlib.Path(p).expanduser())
    except Exception:
        return False   # unreadable / wrong format -> not verified -> fail
    if not isinstance(c14, (int, float)):
        return False   # cell empty / holds text / still a formula string -> fail
    return abs(float(c14) - float(total)) < 1e-9


def assertions():
    """T2's concrete assertions. Both primary end-state checks are MACHINE:
    the artifact file is read directly, nothing falls to intern self-report."""
    return [
        O.file_exists("artifact_path",
                      "the workbook was saved (artifact exists)", primary=True),
        O.Assertion(
            desc="cell C14 equals the exact sum of C2:C13",
            primary=True,
            check=_c14_equals_sum,
            kind=O.MACHINE,
            ctx_key="artifact_path"),
    ]
