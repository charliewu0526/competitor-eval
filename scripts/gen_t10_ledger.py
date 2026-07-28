"""生成 T10 的输入素材 input/ledger.xlsx (月度账本)。

任务: 每月最后一天打开 ledger.xlsx, 重算月度汇总表, 导出 PDF。
故素材含两个 sheet:
  - transactions: 逐笔流水(日期/类别/收支/金额), 供重算。
  - summary: 月度汇总表(留待被重算/刷新的目标 sheet), 初始给出上月旧值,
    让"重算"这个动作有可观测的前后差异。
"""
import pathlib
import openpyxl

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT = ROOT / "tasks" / "T10-excel-schedule-report-001" / "input" / "ledger.xlsx"

wb = openpyxl.Workbook()

tx = wb.active
tx.title = "transactions"
tx.append(["date", "category", "type", "amount"])
rows = [
    ("2025-07-02", "销售收入", "收入", 48000),
    ("2025-07-05", "办公采购", "支出", 3200),
    ("2025-07-08", "销售收入", "收入", 36000),
    ("2025-07-12", "差旅", "支出", 5400),
    ("2025-07-18", "服务收入", "收入", 12500),
    ("2025-07-22", "市场推广", "支出", 8800),
    ("2025-07-28", "工资", "支出", 26000),
]
for r in rows:
    tx.append(r)

# 月度汇总表: 初始留旧值(上月), 等任务在"最后一天"重算刷新成本月。
sm = wb.create_sheet("summary")
sm.append(["月份", "总收入", "总支出", "净额"])
sm.append(["2025-06 (旧值, 待重算为最新月)", 90000, 40000, 50000])

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print("已生成", OUT)
print("transactions 行数:", tx.max_row - 1, " summary 行数:", sm.max_row - 1)
# 供核对的正确汇总(本月): 收入=48000+36000+12500=96500; 支出=3200+5400+8800+26000=43400; 净=53100
inc = sum(a for _, _, t, a in rows if t == "收入")
exp = sum(a for _, _, t, a in rows if t == "支出")
print(f"正确月度汇总(供 expected 参考): 收入={inc} 支出={exp} 净额={inc-exp}")
